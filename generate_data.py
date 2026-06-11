import openpyxl, json
from datetime import datetime

wb = openpyxl.load_workbook('d:/Cline/Agingdashboard/Aging_IMEI_List.xlsx', read_only=True)

def gv(row, key):
    v = row.get(key, '')
    return v if v is not None else ''

def calc_aging_period(days):
    d = int(days)
    if d >= 121: return '121+ Days'
    if d >= 91: return '91~120 Days'
    if d >= 61: return '61~90 Days'
    if d >= 31: return '31~60 Days'
    return 'Under 30 Days'

# Process each sheet as a location (skip "All Locations")
location_sheets = [s for s in wb.sheetnames if s != 'All Locations']

aging_days_arr = []
period_dist = {}
grade_dist = {}
storage_dist = {}
transaction_dist = {}
location_dist = {}
location_aging_cross = {}
location_grade_cross = {}
location_imei_detail = {}
model_map = {}

for sheet_name in location_sheets:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        continue
    header = rows[0]
    raw = [dict(zip(header, row)) for row in rows[1:]]

    loc = sheet_name
    loc_imei_list = []

    for row in raw:
        aging_days = 0
        try:
            aging_days = int(gv(row, 'Aging Days'))
        except (ValueError, TypeError):
            continue
        if aging_days == 0:
            continue

        aging_period = gv(row, 'Aging Period') or calc_aging_period(aging_days)
        grade = gv(row, 'Grade') or '-'
        storage = gv(row, 'Storage') or '-'
        model = gv(row, 'MKT Name') or 'Unknown'
        transaction = gv(row, 'Current Transaction') or '-'
        imei = str(gv(row, 'Final IMEI'))

        aging_days_arr.append(aging_days)
        period_dist[aging_period] = period_dist.get(aging_period, 0) + 1
        grade_dist[grade] = grade_dist.get(grade, 0) + 1
        storage_dist[storage] = storage_dist.get(storage, 0) + 1
        if transaction and transaction != '-':
            transaction_dist[transaction] = transaction_dist.get(transaction, 0) + 1
        location_dist[loc] = location_dist.get(loc, 0) + 1

        if loc not in location_aging_cross:
            location_aging_cross[loc] = {}
        location_aging_cross[loc][aging_period] = location_aging_cross[loc].get(aging_period, 0) + 1

        if loc not in location_grade_cross:
            location_grade_cross[loc] = {}
        location_grade_cross[loc][grade] = location_grade_cross[loc].get(grade, 0) + 1

        loc_imei_list.append({
            'Final IMEI': imei,
            'MKT Name': model,
            'Storage': storage,
            'Grade': grade,
            'Aging Days': aging_days,
            'Aging Period': aging_period
        })

        if model not in model_map:
            model_map[model] = {'name': model, 'count': 0, 'totalAging': 0, 'maxAging': 0, 'over90': 0, 'over120': 0, 'storages': {}, 'grades': {}, 'periodCross': {}}
        m = model_map[model]
        m['count'] += 1
        m['totalAging'] += aging_days
        if aging_days > m['maxAging']:
            m['maxAging'] = aging_days
        if aging_days >= 90:
            m['over90'] += 1
        if aging_days >= 120:
            m['over120'] += 1
        m['storages'][storage] = True
        m['grades'][grade] = True
        m['periodCross'][aging_period] = m['periodCross'].get(aging_period, 0) + 1

    # Sort by aging days descending, keep top 50 for IMEI detail (performance)
    loc_imei_list.sort(key=lambda x: x['Aging Days'], reverse=True)
    location_imei_detail[loc] = loc_imei_list[:50]

wb.close()

total_imei = len(aging_days_arr)
sorted_days = sorted(aging_days_arr)
median = sorted_days[len(sorted_days)//2] if sorted_days else 0
over90 = sum(1 for d in aging_days_arr if d >= 90)
over120 = sum(1 for d in aging_days_arr if d >= 120)

summary = {
    'totalIMEI': total_imei,
    'avgAgingDays': round(sum(aging_days_arr)/total_imei, 1) if total_imei else 0,
    'maxAgingDays': max(aging_days_arr) if aging_days_arr else 0,
    'medianAgingDays': median,
    'over90Days': over90,
    'over120Days': over120,
    'over90Ratio': round(over90/total_imei*1000/10, 1) if total_imei else 0,
    'over120Ratio': round(over120/total_imei*1000/10, 1) if total_imei else 0
}

severity_map = {'31~60 Days': 'Caution (31~60d)', '61~90 Days': 'Warning (61~90d)', '91~120 Days': 'High Risk (91~120d)', '121+ Days': 'Critical (121+d)'}
aging_severity = {severity_map.get(k, k): v for k, v in period_dist.items()}

model_arr = sorted(model_map.values(), key=lambda x: x['count'], reverse=True)

top_models = []
for m in model_arr[:20]:
    top_models.append({
        'name': m['name'],
        'count': m['count'],
        'avgAging': round(m['totalAging']/m['count'], 1),
        'maxAging': m['maxAging'],
        'over90': m['over90'],
        'over90Ratio': round(m['over90']/m['count']*1000/10, 1)
    })

model_aging_cross = []
for m in model_arr[:20]:
    obj = {'name': m['name']}
    for p in ['31~60 Days', '61~90 Days', '91~120 Days', '121+ Days']:
        obj[p] = m['periodCross'].get(p, 0)
    model_aging_cross.append(obj)

all_models = []
for m in model_arr:
    all_models.append({
        'name': m['name'],
        'count': m['count'],
        'storage': ', '.join(sorted(m['storages'].keys())),
        'grades': ', '.join(sorted(m['grades'].keys())),
        'avgAging': round(m['totalAging']/m['count'], 1),
        'maxAging': m['maxAging'],
        'over90': m['over90'],
        'over90Ratio': round(m['over90']/m['count']*1000/10, 1),
        'over120': m['over120']
    })

# Histogram
hist_buckets = {}
for d in aging_days_arr:
    bucket = (d // 15) * 15
    label = f'{bucket}~{bucket+15}d'
    hist_buckets[label] = hist_buckets.get(label, 0) + 1
aging_histogram = [{'range': k, 'count': v} for k, v in sorted(hist_buckets.items(), key=lambda x: int(x[0].split('~')[0]))]

data = {
    'summary': summary,
    'agingPeriodDist': period_dist,
    'agingSeverity': aging_severity,
    'gradeDist': grade_dist,
    'storageDist': storage_dist,
    'transactionDist': transaction_dist,
    'topModels': top_models,
    'modelAgingCross': model_aging_cross,
    'allModels': all_models,
    'agingHistogram': aging_histogram,
    'locationDist': location_dist,
    'locationAgingCross': location_aging_cross,
    'locationGradeCross': location_grade_cross,
    'locationImeiDetail': location_imei_detail,
    'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M')
}

js_content = 'var AGING_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n'
with open('d:/Cline/Agingdashboard/aging_data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f'Generated aging_data.js with {total_imei} records')
print(f'Locations: {list(location_dist.keys())}')
print(f'Location counts: {dict(sorted(location_dist.items(), key=lambda x: -x[1]))}')
print(f'Top model: {top_models[0]["name"]} ({top_models[0]["count"]})')
